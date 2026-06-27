#!/usr/bin/env python3
"""
make_eval_sheets.py — have OpenAI AUTHOR two judging criteria (neutral vs
conversational), then write two blind eval .xlsx sheets over the SAME 30 real
base/soup pairs (eval_pref_hin.csv). GPT writes only the *rubric wording*, never
the translations — the pairs stay real and blind.

Outputs:
  eval_criteria.json       {"neutral": "...", "conversational": "..."}
  eval_neutral_hin.xlsx    30 blind pairs + neutral criterion banner + A/Tie/B dropdown
  eval_conv_hin.xlsx       same pairs + conversational criterion banner

Setup:  pip install openai openpyxl ; export OPENAI_API_KEY=sk-... ; python make_eval_sheets.py
"""
import csv
import json

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

client = OpenAI()
AUTHOR_MODEL = "gpt-4o"      # model that writes the two rubric texts

META = """You are designing a blind A/B evaluation for English->Hindi translation.
A native Hindi speaker will see an English sentence and two Hindi translations
(A and B) and must pick one. Write TWO short judging instructions:

1) "neutral": judge OVERALL translation quality — accuracy of meaning plus
   fluency/grammar. Express no preference for formal vs casual style.

2) "conversational": prioritise natural, CASUAL, spoken/chat-style Hindi; reward
   idiomatic colloquial phrasing, casual pronouns, natural tag-questions and word
   order. BUT make explicit that a genuine error (wrong meaning, broken grammar,
   mistranslation, or a glitch) must still LOSE regardless of how casual it sounds.

Each instruction must contain {en}, {a}, {b} placeholders for the English sentence
and the two translations, and must end by telling the judge to answer with EXACTLY
one token: A, B, or Tie.

Return ONLY JSON: {"neutral": "<instruction>", "conversational": "<instruction>"}."""


def author_criteria():
    resp = client.chat.completions.create(
        model=AUTHOR_MODEL, temperature=0,
        response_format={"type": "json_object"},
        messages=[{"role": "user", "content": META}])
    crit = json.loads(resp.choices[0].message.content)
    assert "neutral" in crit and "conversational" in crit
    for k, v in crit.items():
        assert all(p in v for p in ("{en}", "{a}", "{b}")), f"{k} missing placeholders"
    return crit


def write_sheet(path, banner, rows):
    wb = Workbook(); ws = wb.active; ws.title = "rate"
    # banner row: the criterion, merged across all columns
    ws.append(["CRITERION: " + banner.replace("\n", " ")])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    bc = ws.cell(row=1, column=1)
    bc.font = Font(bold=True, color="9C0006")
    bc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 60
    # header + data
    header = ["id", "english", "translation_A", "translation_B", "which_is_better"]
    ws.append(header)
    for r in rows:
        ws.append([r["id"], r["english"], r["translation_A"],
                   r["translation_B"], ""])
    hdr_fill = PatternFill("solid", fgColor="305496")
    for c in range(1, 6):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A3"
    widths = {"id": 5, "english": 48, "translation_A": 42,
              "translation_B": 42, "which_is_better": 16}
    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = widths[header[c - 1]]
        if header[c - 1] in ("english", "translation_A", "translation_B"):
            for r in range(3, len(rows) + 3):
                ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    dv = DataValidation(type="list", formula1='"A is better,Tie,B is better"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E3:E{len(rows) + 2}")
    wb.save(path)
    print(f"wrote {path}  ({len(rows)} pairs)")


rows = list(csv.DictReader(open("eval_data/eval_pref_hin.csv", encoding="utf-8")))
print(f"authoring criteria with {AUTHOR_MODEL} ...")
crit = author_criteria()
json.dump(crit, open("eval_data/eval_criteria.json", "w", encoding="utf-8"),
          ensure_ascii=False, indent=2)
print("wrote eval_criteria.json\n")
for k in ("neutral", "conversational"):
    print(f"--- {k} criterion (GPT-authored) ---\n{crit[k]}\n")
write_sheet("eval_data/eval_neutral_hin.xlsx", crit["neutral"], rows)
write_sheet("eval_data/eval_conv_hin.xlsx", crit["conversational"], rows)
print("\nnext: python dual_criterion_judge.py    (runs top-10 models on both)")
