#!/usr/bin/env python3
"""Convert the human-eval CSVs to rater-friendly .xlsx (dropdowns, wrap, widths)."""
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def make_sheet(csv_path, xlsx_path, score_cols):
    rows = list(csv.reader(open(csv_path, encoding="utf-8")))
    wb = Workbook()
    ws = wb.active
    ws.title = "rate"
    for r in rows:
        ws.append(r)
    # header style
    hdr_fill = PatternFill("solid", fgColor="305496")
    for c in range(1, len(rows[0]) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    # column widths + wrap for text columns
    widths = {"id": 5, "english": 50, "reference": 45,
              "translation_A": 45, "translation_B": 45}
    for c in range(1, len(rows[0]) + 1):
        name = rows[0][c - 1]
        col = get_column_letter(c)
        ws.column_dimensions[col].width = widths.get(name, 14)
        if name in ("english", "reference", "translation_A", "translation_B"):
            for r in range(2, len(rows) + 1):
                ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    # 1-5 dropdown on score columns
    if score_cols:
        dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
        ws.add_data_validation(dv)
        for c in score_cols:
            col = get_column_letter(c)
            dv.add(f"{col}2:{col}{len(rows)}")
    wb.save(xlsx_path)
    print(f"{csv_path} -> {xlsx_path}  ({len(rows)-1} rows)")


# eval sheets: score columns F,G,H,I = 6,7,8,9
make_sheet("eval_data/eval_sheet_hin.csv", "eval_data/eval_sheet_hin.xlsx", [6, 7, 8, 9])
make_sheet("eval_data/eval_sheet_tam.csv", "eval_data/eval_sheet_tam.xlsx", [6, 7, 8, 9])
# keys: no dropdowns (keep private)
make_sheet("eval_data/eval_key_hin.csv", "eval_data/eval_key_hin.xlsx", [])
make_sheet("eval_data/eval_key_tam.csv", "eval_data/eval_key_tam.xlsx", [])
