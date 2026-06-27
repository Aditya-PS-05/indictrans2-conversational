#!/usr/bin/env python3
"""
analyze_eval.py SHEET.csv KEY.csv  (pure stdlib)
Un-blind a filled human-eval sheet (join A/B scores to base/soup via the key),
report base vs. soup mean adequacy & fluency, and a paired bootstrap p-value.
Multiple raters: pass several SHEET.csv files before the KEY (same key).
  python3 analyze_eval.py eval_sheet_hin.csv eval_key_hin.csv
  python3 analyze_eval.py rater1_hin.csv rater2_hin.csv eval_key_hin.csv
"""
import csv
import random
import statistics
import sys

random.seed(0)
*sheets, key_path = sys.argv[1:]


def read_dicts(path):
    """Read rows as dicts from a .csv or .xlsx file."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        rows = list(load_workbook(path, data_only=True).active.iter_rows(values_only=True))
        header = [str(h) if h is not None else "" for h in rows[0]]
        for r in rows[1:]:
            yield {header[i]: ("" if i >= len(r) or r[i] is None else r[i])
                   for i in range(len(header))}
    else:
        with open(path, encoding="utf-8") as f:
            yield from csv.DictReader(f)


key = {}  # id -> (A_system, B_system)
for row in read_dicts(key_path):
    key[str(row["id"])] = (row["A_system"], row["B_system"])

# per-sentence scores mapped to system
adq = {"base": [], "soup": []}
flu = {"base": [], "soup": []}
# keep paired arrays (soup - base) per sentence for the bootstrap
pair_adq, pair_flu = [], []
skipped = 0

for sheet in sheets:
    for row in read_dicts(sheet):
            rid = str(row["id"])
            if rid not in key:
                continue
            try:
                aa, af = float(row["A_adequacy_1to5"]), float(row["A_fluency_1to5"])
                ba, bf = float(row["B_adequacy_1to5"]), float(row["B_fluency_1to5"])
            except (ValueError, KeyError):
                skipped += 1
                continue
            a_sys, b_sys = key[rid]
            s_adq = {a_sys: aa, b_sys: ba}
            s_flu = {a_sys: af, b_sys: bf}
            for sysname in ("base", "soup"):
                adq[sysname].append(s_adq[sysname])
                flu[sysname].append(s_flu[sysname])
            pair_adq.append(s_adq["soup"] - s_adq["base"])
            pair_flu.append(s_flu["soup"] - s_flu["base"])

n = len(pair_adq)
if n == 0:
    print("No scored rows found. Did you fill in the score columns?")
    sys.exit(1)


def boot_p(diffs, iters=10000):
    """two-sided paired bootstrap p-value that mean diff != 0."""
    m = statistics.mean(diffs)
    cnt = 0
    for _ in range(iters):
        sample = [diffs[random.randrange(n)] for _ in range(n)]
        if (statistics.mean(sample) <= 0) == (m > 0):
            cnt += 1
    return 2 * cnt / iters  # two-sided


print(f"Scored sentences: {n}  (skipped/unfilled: {skipped})  raters: {len(sheets)}")
print(f"{'metric':10s} {'base':>6s} {'soup':>6s} {'Δ':>6s} {'p (boot)':>9s}")
for name, sc, pairs in [("adequacy", adq, pair_adq), ("fluency", flu, pair_flu)]:
    b, s = statistics.mean(sc["base"]), statistics.mean(sc["soup"])
    print(f"{name:10s} {b:>6.2f} {s:>6.2f} {s-b:>+6.2f} {boot_p(pairs):>9.4f}")
