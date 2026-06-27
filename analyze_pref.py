#!/usr/bin/env python3
"""
analyze_pref.py SHEET(.csv/.xlsx) KEY.csv
Un-blind a filled pairwise-preference sheet and count base vs soup wins, with a
two-sided binomial sign test (ties excluded). Multiple raters: list sheets first.
"""
import csv
import math
import sys

*sheets, key_path = sys.argv[1:]


def read_dicts(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        rows = list(load_workbook(path, data_only=True).active.iter_rows(values_only=True))
        header = [str(h) if h is not None else "" for h in rows[0]]
        for r in rows[1:]:
            yield {header[i]: ("" if i >= len(r) or r[i] is None else r[i])
                   for i in range(len(header))}
    else:
        with open(path, encoding="utf-8") as f:
            yield from csv.DictReader(f)


key = {}
for row in read_dicts(key_path):
    key[str(row["id"])] = (row["A_system"], row["B_system"])

soup_wins = base_wins = ties = blank = 0
for sheet in sheets:
    for row in read_dicts(sheet):
        rid = str(row["id"])
        if rid not in key:
            continue
        choice = str(row.get("which_is_better", "")).strip().lower()
        a_sys, b_sys = key[rid]
        if choice.startswith("a"):
            winner = a_sys
        elif choice.startswith("b"):
            winner = b_sys
        elif choice.startswith("t"):
            ties += 1; continue
        else:
            blank += 1; continue
        if winner == "soup":
            soup_wins += 1
        else:
            base_wins += 1


def binom_two_sided(k, n, p=0.5):
    if n == 0:
        return 1.0
    obs = math.comb(n, k) * p**k * (1 - p)**(n - k)
    tot = 0.0
    for i in range(n + 1):
        pi = math.comb(n, i) * p**i * (1 - p)**(n - i)
        if pi <= obs + 1e-12:
            tot += pi
    return min(1.0, tot)


decisive = soup_wins + base_wins
p = binom_two_sided(soup_wins, decisive)
print(f"raters: {len(sheets)}  |  decisive: {decisive}  ties: {ties}  blank: {blank}")
print(f"soup preferred: {soup_wins}   base preferred: {base_wins}   (ties: {ties})")
if decisive:
    print(f"soup win-rate (excl. ties): {soup_wins/decisive:.0%}   sign-test p = {p:.4f}"
          + ("  *significant*" if p < 0.05 else ""))
